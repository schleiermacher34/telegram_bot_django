# botapp/utils.py

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from .models import UserData, VideoLink
from django.http import HttpResponse

def export_data_to_excel():
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Export UserData
    ws_users = wb.create_sheet(title='UserData')
    
    # Define the headers
    headers_users = ['Telegram Username', 'Chat ID', 'Serial Number']
    ws_users.append(headers_users)
    
    # Fetch all UserData entries
    users = UserData.objects.all()
    
    # Write data rows
    for user in users:
        row = [
            user.telegram_username or '',
            user.chat_id,
            user.serial_number,
        ]
        ws_users.append(row)
    
    # Adjust column widths
    for col_num, _ in enumerate(headers_users, 1):
        column_letter = get_column_letter(col_num)
        ws_users.column_dimensions[column_letter].width = 20
    
    # Export VideoLinks
    ws_videos = wb.create_sheet(title='VideoLinks')
    
    # Define the headers
    headers_videos = ['Lottery Ticket Number', 'Telegram Username', 'Chat ID', 'Serial Number', 'Video Link']
    ws_videos.append(headers_videos)
    
    # Fetch all VideoLink entries with related UserData
    video_links = VideoLink.objects.select_related('user').all()
    
    # Write data rows
    for video in video_links:
        row = [
            video.id,
            video.user.telegram_username or '',
            video.user.chat_id,
            video.user.serial_number,
            video.video_link,
        ]
        ws_videos.append(row)
    
    # Adjust column widths
    for col_num, _ in enumerate(headers_videos, 1):
        column_letter = get_column_letter(col_num)
        ws_videos.column_dimensions[column_letter].width = 25
    
    # Save the workbook to a response
    response = HttpResponse(content_type='application/ms-excel')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'exported_data_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
